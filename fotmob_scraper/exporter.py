"""파싱된 FotMob 데이터를 Excel로 내보내는 모듈."""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1A1A2E")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL = PatternFill("solid", fgColor="F0F4FF")
BORDER = Border(
    bottom=Side(style="thin", color="CCCCCC"),
)

SHEET_ORDER = ["개요", "통계", "이벤트", "라인업", "슈팅"]


def _apply_header(ws, headers: list[str]):
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val.encode("cp949", errors="replace")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)


def _write_sheet_from_list(ws, rows: list[dict]):
    if not rows:
        ws.append(["데이터 없음"])
        return
    headers = list(rows[0].keys())
    _apply_header(ws, headers)
    for i, row in enumerate(rows, start=2):
        ws.append([row.get(h) for h in headers])
        if i % 2 == 0:
            for cell in ws[i]:
                cell.fill = ALT_FILL
    _auto_width(ws)


def export_to_excel(parsed: dict, output_path: str) -> str:
    """
    파싱된 데이터를 Excel 파일로 저장.

    Args:
        parsed: parse_all() 결과 dict
        output_path: 저장할 파일 경로 (.xlsx)

    Returns:
        저장된 파일의 절대 경로
    """
    wb = Workbook()
    wb.remove(wb.active)  # 기본 Sheet 제거

    # ── 개요 시트 ──────────────────────────────
    ws_overview = wb.create_sheet("개요")
    overview = parsed.get("overview", {})
    ws_overview.append(["항목", "값"])
    for cell in ws_overview[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for k, v in overview.items():
        ws_overview.append([k, str(v) if v is not None else ""])
    _auto_width(ws_overview)

    # ── 통계 시트 ──────────────────────────────
    ws_stats = wb.create_sheet("통계")
    _write_sheet_from_list(ws_stats, parsed.get("stats", []))

    # ── 이벤트 시트 ────────────────────────────
    ws_events = wb.create_sheet("이벤트")
    _write_sheet_from_list(ws_events, parsed.get("events", []))

    # ── 라인업 시트 ────────────────────────────
    ws_lineup = wb.create_sheet("라인업")
    _write_sheet_from_list(ws_lineup, parsed.get("lineups", []))

    # ── 슈팅 시트 ──────────────────────────────
    ws_shots = wb.create_sheet("슈팅")
    _write_sheet_from_list(ws_shots, parsed.get("shots", []))

    path = Path(output_path).resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return str(path)


def save_raw_json(data: dict, output_path: str) -> str:
    """원본 API JSON 데이터를 파일로 저장 (디버깅용)."""
    path = Path(output_path).resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return str(path)
